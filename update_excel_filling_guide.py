"""
更新 Excel 填寫範例：在標題列（第 1 列）下方寫入填寫注意事項（第 2 列），
並將可為 null 的欄位在第 1、2 列的儲存格設為灰色背景。
不變更欄位順序、欄位名稱與資料列內容。
"""
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 灰色背景（可選欄位），與 create_new_excel_template 的 optional_fill 一致
GREY_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# 各工作表「可為 null」的欄位（其餘為必要，不塗灰）
OPTIONAL_FIELDS = {
    "UI_SEGMENTS": {"order", "mode", "tag", "published", "configId", "topicIds"},
    "TOPICS": {"published", "order", "tags", "bubbleImageUrl", "bubbleStorageFile", "bubbleGradStart", "bubbleGradEnd", "createdAt", "updatedAt"},
    "PRODUCTS": {
        "title", "titleLower", "order", "type", "published", "levelGoal", "levelBenefit",
        "anchorGroup", "version", "coverImageUrl", "coverStorageFile", "itemCount", "wordCountAvg",
        "pushStrategy", "sourceType", "source", "sourceUrl",
        "spec1Label", "spec2Label", "spec3Label", "spec4Label",
        "spec1Icon", "spec2Icon", "spec3Icon", "spec4Icon",
        "trialMode", "trialLimit", "releaseAtMs", "createdAtMs",
        "contentarchitecture",
    },
    "FEATURED_LISTS": {"type", "topicIds", "productIds", "published", "order", "updatedAt", "ids", "coverImageUrl", "coverStorageFile"},
    "CONTENT_ITEMS": {
        "type", "topicId", "level", "levelGoal", "levelBenefit", "anchorGroup", "anchor",
        "intent", "difficulty", "content", "wordCount", "reusable",
        "sourceType", "source", "sourceUrl", "version", "pushOrder", "storageFile",
        "seq", "isPreview", "deepAnalysis", "mediaImageUrl", "mediaStorageFile",
    },
}

# 第 2 列「填寫注意事項」文案（依欄位名）
ROW2_HINTS = {
    "UI_SEGMENTS": {
        "configId": "設定 ID（選填）",
        "segmentId": "區段 ID（唯一識別碼）",
        "title": "區段標題",
        "order": "排序數字，越小越前；預設 0",
        "mode": "all 或 tag",
        "tag": "標籤（mode=tag 時用）；選填",
        "topicIds": "主題 ID 列表（選填）",
        "published": "true/false；選填",
    },
    "TOPICS": {
        "topicId": "主題 ID（唯一識別碼）",
        "title": "主題標題",
        "published": "true/false；選填",
        "order": "排序數字；選填",
        "tags": "標籤，多個用分號 ; 分隔",
        "bubbleImageUrl": "泡泡圖片 URL（選填，建議用 bubbleStorageFile）",
        "bubbleStorageFile": "泡泡圖 Storage 路徑。圖片：512×512 px, PNG, <100KB",
        "bubbleGradStart": "漸層起始色 HEX，例 #FF5733",
        "bubbleGradEnd": "漸層結束色 HEX，例 #33FF57",
        "createdAt": "建立時間；選填",
        "updatedAt": "更新時間；選填",
    },
    "PRODUCTS": {
        "productId": "產品 ID（唯一識別碼）",
        "type": "類型，如 course；選填",
        "topicId": "所屬主題 ID",
        "level": "等級，如 L1、L2",
        "levelGoal": "等級目標描述；選填",
        "levelBenefit": "等級效益描述；選填",
        "anchorGroup": "錨點群組；選填",
        "version": "版本號；選填",
        "title": "產品標題；留空則自動用 topicId+level",
        "titleLower": "標題小寫；留空則自動生成；選填",
        "order": "排序，預設 0；選填",
        "published": "true/false；選填",
        "coverImageUrl": "封面圖 URL；選填",
        "coverStorageFile": "封面圖 Storage 路徑。圖片：1200×800 px, JPG/PNG, <300KB",
        "itemCount": "內容數量（數字）；選填",
        "wordCountAvg": "平均字數；選填",
        "pushStrategy": "推播策略；選填",
        "sourceType": "來源類型；選填",
        "source": "來源；選填",
        "sourceUrl": "來源 URL；選填",
        "spec1Label": "規格 1 標籤；選填",
        "spec2Label": "規格 2 標籤；選填",
        "spec3Label": "規格 3 標籤；選填",
        "spec4Label": "規格 4 標籤；選填",
        "spec1Icon": "Material Icons 名稱，如 timer；選填",
        "spec2Icon": "Material Icons 名稱；選填",
        "spec3Icon": "Material Icons 名稱；選填",
        "spec4Icon": "Material Icons 名稱；選填",
        "trialMode": "試用模式；選填",
        "trialLimit": "試用數量上限，預設 3；選填",
        "releaseAtMs": "發布時間戳（毫秒）；選填",
        "createdAtMs": "建立時間戳（毫秒）；選填",
        "contentarchitecture": "內容架構說明；選填；產品頁「內容架構」卡片顯示",
    },
    "FEATURED_LISTS": {
        "listId": "清單 ID（唯一識別碼）",
        "title": "清單標題",
        "type": "productIds 或 topicIds；選填預設 productIds",
        "topicIds": "主題 ID 列表，分號 ; 分隔；選填",
        "productIds": "產品 ID 列表，分號 ; 分隔；選填",
        "published": "true/false；選填",
        "order": "排序；選填",
        "updatedAt": "更新時間；選填",
        "ids": "ID 列表（分號分隔）；選填",
        "coverImageUrl": "精選清單封面圖 URL。1200×800 px, JPG/PNG, <300KB；選填",
        "coverStorageFile": "精選清單封面圖 Storage 路徑；規格同上；選填",
    },
    "CONTENT_ITEMS": {
        "itemId": "內容項目 ID（唯一識別碼）",
        "productId": "所屬產品 ID",
        "type": "內容類型，如 card、sentence、quote；選填",
        "topicId": "所屬主題 ID；選填",
        "level": "等級；選填",
        "levelGoal": "等級目標；選填",
        "levelBenefit": "等級效益；選填",
        "anchorGroup": "錨點群組；選填",
        "anchor": "錨點；選填",
        "intent": "建議：tips / learn / practice / review 等小寫關鍵字",
        "difficulty": "1=最簡單 2=簡單 3=中等 4=難 5=最難（預設 1）",
        "content": "內容文字；選填",
        "wordCount": "字數（數字）；選填",
        "reusable": "true/false；選填",
        "sourceType": "來源類型；選填",
        "source": "來源；選填",
        "sourceUrl": "來源 URL；選填",
        "version": "版本號；選填",
        "pushOrder": "推播順序（數字，Day N）；選填",
        "storageFile": "Storage 路徑。圖片最大寬 2000px, PNG/JPG <1MB；PDF <5MB",
        "seq": "序列號，預設 0；選填",
        "isPreview": "true/false；選填",
        "deepAnalysis": "深度分析內文；選填",
        "mediaImageUrl": "媒體圖片 URL；選填",
        "mediaStorageFile": "媒體圖 Storage 路徑；選填",
    },
}


def get_row1_headers(ws):
    """取得第 1 列欄位名稱列表（依實際儲存格值）。"""
    headers = []
    col = 1
    while True:
        cell = ws.cell(row=1, column=col)
        val = cell.value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            break
        headers.append(str(val).strip() if isinstance(val, str) else str(val))
        col += 1
    return headers


def apply_filling_guide(wb, sheet_name):
    """對指定工作表：插入第 2 列、寫入注意事項、可選欄位塗灰。"""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    headers = get_row1_headers(ws)
    if not headers:
        return

    optional = OPTIONAL_FIELDS.get(sheet_name, set())
    hints = ROW2_HINTS.get(sheet_name, {})

    # 插入一列於第 2 列（原第 2 列起之資料下移）
    ws.insert_rows(2)

    for col_idx, col_name in enumerate(headers, start=1):
        # 第 2 列：填寫注意事項
        hint = hints.get(col_name, "選填" if col_name in optional else "必填")
        cell_r2 = ws.cell(row=2, column=col_idx)
        cell_r2.value = hint
        cell_r2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 可選欄位：第 1、2 列塗灰
        if col_name in optional:
            ws.cell(row=1, column=col_idx).fill = GREY_FILL
            cell_r2.fill = GREY_FILL

    # 設定第 2 列行高以便顯示多行說明
    if ws.row_dimensions[2].height is None or ws.row_dimensions[2].height < 30:
        ws.row_dimensions[2].height = 40


def main():
    ap = argparse.ArgumentParser(description="更新 Excel 填寫範例：加入第 2 列注意事項與可選欄位灰色標示")
    ap.add_argument("excel", help="Excel 檔案路徑（如 excel填寫範例.xlsx）")
    ap.add_argument("--output", "-o", default=None, help="輸出路徑；未指定則覆寫原檔")
    args = ap.parse_args()

    input_path = args.excel
    output_path = args.output or input_path

    wb = load_workbook(input_path)
    for sheet_name in list(wb.sheetnames):
        apply_filling_guide(wb, sheet_name)
        print(f"  已處理工作表: {sheet_name}")

    wb.save(output_path)
    print(f"\n已儲存: {output_path}")


if __name__ == "__main__":
    main()
